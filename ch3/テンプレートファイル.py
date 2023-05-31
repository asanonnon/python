import openpyxl as excel

## テンプレートファイルから請求書を作成
# ファイル名を指定
template_file = "ch3/invoice-template.xlsx" # テンプレートファイル
save_file = "excel/invoice01.xlsx"          # セーブファイル
# 設定するデータ
name = "田中一郎"
subject = "1月分のご請求"
items =[
    ["リンゴ",5,320],
    ["バナナ",8,210],
    ["マロン",1,1200]
]

# テンプレートを開く
book = excel.load_workbook(template_file)
sheet = book.active
# テンプレートに名前と件名を書き込む
sheet["B4"] = name
sheet["C10"] = subject
# 内訳を連続で書き込む
total = 0 
for i,it in enumerate(items):
    summary,count,price = it #　配列を分解
    subtotal = count * price
    total += subtotal
    # シートに書き込む
    row = 15 + i
    sheet.cell(row,2,summary)
    sheet.cell(row,5,count)
    sheet.cell(row,6,price)
    sheet.cell(row,7,subtotal)
# 請求金額（合計金額）を書き込む
sheet["C11"] = total
book.save(save_file) # saveの時に違うファイルを指定することで元のファイルを崩さずに再利用できる