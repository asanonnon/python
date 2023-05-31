import openpyxl as excel

book = excel.load_workbook("all-customer.xlsx")       
sheet = book["名簿"]

########################## 顧客一覧を確認してシートに分ける##############################

for row in sheet.iter_rows(min_row=3):              
    cells = [v.value for v in row]                  # 一つずつ要素を取得する
    if cells[0] is None: break
    print(cells)

    (name,area,plan) = cells                        # 各列のデータを変数に代入

    sname = plan+"プラン"                            # コピー先のシート名を決める

    ####  シートの新規作成または指定して、取ってきた要素を追加する  #####

    if sname not in book.sheetnames:                # 該当シートがあるか
        to_sheet = book.create_sheet(title=sname)   # ない場合新しく作り、配列[0]「見出し行」を追加する
        to_sheet.append(["名前","住所","プラン"])
    else:
        to_sheet = book[sname]                      # ある場合to_sheetをsnameを指定

    to_sheet.append(cells)                          #　該当シートに顧客情報を追記　※シートに対してappendできる

book.save("split_sheet.xlsx")