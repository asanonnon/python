import openpyxl as xl 
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border,Side
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

book = xl.Workbook()
sheet = book.active

# 横幅を設定
sheet.column_dimensions["B"].width = 40
# 高さを設定
sheet.row_dimensions[2].height = 40

cell = sheet["B2"]
cell.value = "喜びにあふれた心はいい薬"

# テキストの配置
cell.alignment = Alignment(
    horizontal = "center",
    vertical = "center"
)


# 罫線の指定
cell.border = Border(
    top=Side(style="thin",color="000000"),
    right=Side(style="thin",color="000000"),
    bottom=Side(style="thin",color="000000"),
    left=Side(style="thin",color="000000"),
)

# フォントの指定
cell.font = Font(
    size=14,
    bold=True,
    italic=True,
    color="FFFFFF"
)

#　背景色の指定
cell.fill = PatternFill(
    fill_type="solid",
    fgColor="FF0000"
)

book.save('style.xlsx')

