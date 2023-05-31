########################### 複数のブックとシートを操作する###########################

# import openpyxl as excel 　                                  openpyxlを取り込む 

# book = excel.Workbook()　                                    新規ワークブックを作成

# book  = excel.load_workbook('ファイル名.xlsx')　              既存のワークシートをファイルから開く

# dook = excel.load_workbook('ファイル名.xlsx',data_only=True)　ワークブックを開く（式があれば展開して開く）

# book.close()　ワークブックを明示的に閉じる


############################ブックから任意のシートを得る方法########################

# sheet = book.active　                                        アクティブなシートを得る

# sheet = book.worksheet[シート番号]                           任意の箇所にあるワークシートを得る（0起点）

# sheet = book["シート名"]                                     シート名（sheet1など）を指定して取得

# print(book.sheetnames)                                      ブック内のシート名の一覧を得る

#####################################シートの扱い方の方法##########################

# sheet = bool.create_sheet(title="シート名")                  新規シートを作成
        
# sheet = book.copy_worksheet(book['シート名'])　              既存のシートをコピーして得る
        
# sheet.title = "新しい名前"　                                 シート名を変更する
        
# book.remove(book["シート名"])                                シートの削除

#################################################################################

import openpyxl as excel
import random
# あたりシートの番号を決める
atari = random.randint(1,100)

# 新規ブックの作成
book = excel.Workbook()
book.active["B2"] = "あたりが書かれたシートを探そう"

# 繰り返し100回シートを作成する
for i in range(1,101):
    #　新規シート作成
    sname = str(i) + "番"
    sheet = book.create_sheet(title=sname)
    #　シートに掻っ込む単語を決定
    word ="ハズレ"
    if i == atari: word = "あたり"
    # セルに書き込む
    for y in range(50):
        for x in range(30):
            c = sheet.cell(y+1,x+1)
            c.value = word

book.save("game100.xlsx")
print("ok ,atari=",atari)