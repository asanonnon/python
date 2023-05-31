import openpyxl as excel

book = excel.load_workbook("all-customer.xlsx") # 顧客一覧のブックを開く

sheet = book["名簿"]                            # 名簿シートの選択

customers = [["名前","住所","購入プラン"]]       # 抜き出す顧客を記録する変数

for row in sheet.iter_rows(min_row=3):         # 顧客一覧の抽出 3行目から下すべてを取ってくる
    values = [v.value for v in row]            # リスト内包表記　一覧の数だけリスト化される
    if values[0] is None: break                # values[0]の場合for文を中止する

    area = values[1]                           # 名古屋と横浜ならコピー
    if area == "横浜市" or area == "名古屋市":  # 仮に作ったリストを参照して横浜か名古屋だったらcustomersに２次元配列として追加する
        customers.append(values)
        print(values)

new_book = excel.Workbook()                    # 新規ブックの作成
new_sheet = new_book.active                    # ブックの指定　※最後に開いたファイル
new_sheet["A1"] = "横浜と名古屋の顧客名簿"

for row,row_val in enumerate(customers):       # 抽出したデータを繰り返しシートに書き込む row はインデックス番号　row_valは２次元配列の中の配列　customersは２次元配列
    for col,val in enumerate(row_val):         # row_val は配列　colはインデックス番号　valは配列の中身
        c = new_sheet.cell(2+row,1+col)
        c.value = val

new_book.save('yokohama_nagoya.xlsx')          # ファイルに書き込む