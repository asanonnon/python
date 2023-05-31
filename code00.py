import openpyxl as excel

#　売上データのブックを開いてシートを取り出す
sheet = excel.load_workbook(
    #　data_only=Trueでエクセルで取ってきたデータの計算式を実行した後の結果を取ってくる
    "uriage.xlsx",data_only=True
).active

#　iter_rowsを使って全データを取り出す
for row in sheet.iter_rows(min_row=3):
    # リスト内包表記
    values = [cell.value for cell in row]
    if values[0] is None: break
    print(values)